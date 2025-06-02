import os
import requests
from tkinter import Tk, Button, Label, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
from tkinter import font

def formatar_planilha(wb):
    ws = wb.active

    # 1) Cabeçalho em negrito, cor de fundo e centralizado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # tom de azul
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 2) Ajuste de largura das colunas
    dims = {
        "A": 20,  # Data e Hora
        "B": 15,  # Cidade
        "C": 18,  # Temperatura
        "D": 25,  # Descrição
        "E": 12,  # Umidade
    }
    for col, width in dims.items():
        ws.column_dimensions[col].width = width

    # 3) Bordas finas ao redor de todo o bloco de dados
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border

    # 4) Congelar a primeira linha (cabeçalho)
    ws.freeze_panes = "A2"

def buscar_previsao():
    try:
        cidade = "São Paulo"
        API_KEY = "9ce5a7793761ebb6ade5f6fac171a236"
        link = (
            f"https://api.openweathermap.org/data/2.5/weather"
            f"?q={cidade}&appid={API_KEY}&units=metric&lang=pt_br"
        )
        requisicao = requests.get(link, verify=False)
        dados = requisicao.json()

        temperatura = dados["main"]["temp"]
        descricao = dados["weather"][0]["description"]
        umidade = dados["main"]["humidity"]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        caminho = os.path.abspath("clima_atual.xlsx")

        # Se não existir, cria com cabeçalho
        if not os.path.exists(caminho):
            wb = Workbook()
            ws = wb.active
            ws.append(
                ["Data e Hora", "Cidade", "Temperatura (°C)", "Descrição", "Umidade (%)"]
            )
            formatar_planilha(wb)
            wb.save(caminho)

        # Carrega e adiciona nova linha
        wb = load_workbook(caminho)
        ws = wb.active
        ws.append([timestamp, cidade, temperatura, descricao, umidade])

        # Reaplica o estilo (para que as novas linhas ganhem bordas, etc.)
        formatar_planilha(wb)
        wb.save(caminho)

        messagebox.showinfo(
            "Sucesso",
            (
                f"Dados salvos em {timestamp}\n\n"
                f"Temperatura: {temperatura}°C\n"
                f"Condição: {descricao}\n"
                f"Umidade: {umidade}%"
            ),
        )

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")

# Configuração da interface gráfica com Tkinter
root = Tk()
root.title("Captador de Temperatura")
root.geometry("400x250")  # Tamanho da janela
root.config(bg="#F0F8FF")  # Cor de fundo azul claro

# Estilo do título
titulo_font = font.Font(family="Helvetica", size=18, weight="bold")
titulo_label = Label(
    root, text="Captador de Temperatura", font=titulo_font, bg="#F0F8FF", fg="#4F81BD"
)
titulo_label.pack(pady=20)

# Função para mudar cor ao passar o mouse (hover effect)
def on_enter(event):
    botao_buscar.config(bg="#B8D5E3")

def on_leave(event):
    botao_buscar.config(bg="#4F81BD")

# Botão que chama a função de captura dos dados
botao_buscar = Button(
    root,
    text="Buscar Previsão",
    command=buscar_previsao,
    width=20,
    height=2,
    bg="#4F81BD",
    fg="white",
    font=("Helvetica", 12, "bold"),
    relief="flat",
    bd=5,
    activebackground="#B8D5E3",
)
botao_buscar.pack(pady=30)

# Efeitos de hover no botão
botao_buscar.bind("<Enter>", on_enter)
botao_buscar.bind("<Leave>", on_leave)

# Inicia a interface
root.mainloop()
